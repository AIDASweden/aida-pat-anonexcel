#!/usr/bin/env python3
"""AIDA Pathology Anonymizer for Excel Sheets v%s

Research software not approved for clinical use.

Requirements:
* anonymize_wsi.exe
* Python 3 with the following packages:
    * openpyxl
"""
__version__ = "1.0.0"
__doc__ %= __version__

import argparse
import os
import re
import shutil
import subprocess
import sys
import zipfile

import openpyxl

options = {
    '--version': dict(
        action='version',
        version='%(prog)s ' + __version__),
    'excelfile': dict(
        type=argparse.FileType('rb+'),
        metavar="EXCELFILE.xslx",
        help="An AIDA Pathology Anonymization Excel Sheet. Will be updated in place"),
    '--tmpdir': dict(
        metavar="DIR",
        help="Where to keep temporary files. Default: EXCELFILE_tmp."),
    '--anondir': dict(
        metavar="DIR",
        help="Where to save anonymized slides. Default: EXCELFILE_tmp."),
    ('--suppress-exitcode','-z'): dict(
        action='store_true',
        help="Do not use an exitcode when exiting. Only useful with py -i."),
}
def get_options(argv):
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    for name, params in options.items():
        if isinstance(name, str):
            parser.add_argument(name, **params)
        else:
            parser.add_argument(*name, **params)
    return parser.parse_args(argv[1:])

slidefiletypes = ['.svs', '.ndpi']

tests = dict(
    D1="AIDA Pathology Anonymization Sheet",
    A12="Status",
    B12="Case",
    C12="OrigFile",
    D12="AnonID",
    E12="Block",
    F12="Stain",
)
def check_worksheet(ws):
    for cell, value in tests.items():
        if ws[cell].value != value:
            row = int(cell[1:])
            err(row, "Unsupported excelfile format. Cell {!r} should be {!r} but is {!r}.",
                cell, value, ws[cell].value)

class ParseError(ValueError):
    """Errors encountered while parsing/anonymizing data from Excel spreadsheet."""

def err(row, msg, *args, **kw):
    raise ParseError(row, msg.format(*args, **kw))

def validate_anonid_number(row, anonid, prefix, digits, previous):
    try:
        anonid_number = int(anonid[len(prefix):])
    except:
        anonid_number = -1
    if not (anonid.startswith(prefix) and
            anonid_number > 0 and
            len(anonid) == len(prefix) + digits):
            anonid_number = -1
    if anonid_number < 0:
        err(row, "AnonID is {!r}, but must start with Prefix {!r} followed by a {}-digit number > 0.",
            anonid, prefix, digits)
    if anonid_number < previous:
        err(row, "AnonID number is {!r}, but must not be less than the previous ({}).",
            anonid_number, previous)
    return anonid_number

def validate_id_mapping(row, caseid, anonid, caseids, anonids):
    existing_caseid = caseids.setdefault(anonid, caseid)
    if existing_caseid and existing_caseid != caseid:
        err(row, "Case {} is given AnonID {} on this row, but this AnonID was previously given to Case {}.",
            caseid, anonid, existing_caseid)
    existing_anonid = anonids.setdefault(caseid, anonid)
    if existing_anonid and existing_anonid != anonid:
        err(row, "Case {} is given AnonID {} on this row, but this Case was previously given AnonID {}.",
            caseid, anonid, existing_anonid)

def get_barcode(anonid, block, stain):
    return "{0};{0};{1};{2}".format(anonid, block, stain)

def files(directory, *ext):
    if ext:
        return [f for f in next(os.walk(directory))[2] if os.path.splitext(f)[1] in ext]
    return next(os.walk(directory))[2]

def subdirs(directory):
    return next(os.walk(directory))[1]

def get_slides(i, casedir):
    slides = []
    for slidedir in subdirs(casedir):
        if '_' not in slidedir:
            err(i, "Slide directory {!r} for Case {!r} is not named on the format BLOCK_STAIN (eg 'A_HE').",
                slidedir, os.path.basename(casedir))
        block, stain = slidedir.split('_')
        slidefiles = files(os.path.join(casedir, slidedir), *slidefiletypes)
        if len(slidefiles) != 1:
            err(i, "Expected 1 slide file in slide directory {!r} for Case {!r} but found {}{}.",
                slidedir, os.path.basename(casedir), len(slidefiles),
                ' ({})'.format(', '.join(repr(f) for f in slidefiles) if slidefiles else ''))
        origfile = os.path.join(casedir, slidedir, slidefiles[0])
        slides.append((block, stain, origfile))
    return slides

def anonymize_slide(i, origfile, anondir, barcode):
    anonymize_cmd = ["anonymize_wsi", "-o", anondir, "-bv", barcode, origfile]
    try:
        subprocess.run(anonymize_cmd, check=True)
    except:
        err(i, "anonymize_wsi returned a nonzero exit code for command {}. Aborting.",
            " ".join(repr(arg) for arg in anonymize_cmd))

def mark_red(cell):
    cell.fill = openpyxl.styles.PatternFill(start_color="FFC7CE", fill_type = "solid")
    cell.font = openpyxl.styles.Font(color="FF0000")

def mark_ok(cell):
    cell.style = 'Normal'

def mark_done(worksheet, row, case, origfile):
    p = worksheet.cell(row, 2)
    o = worksheet.cell(row, 3)
    mark_red(p) if case else mark_ok(p)
    mark_red(o) if origfile else mark_ok(o)

def update_spreadsheet(i, ws, case, origfile, anonid, slides):
    ws.insert_rows(i+1, len(slides) - 1)
    for slideindex, (block, stain, origfile) in enumerate(slides):
        barcode = get_barcode(anonid, block, stain)
        ws.cell(i + slideindex, 1, 'Done')
        ws.cell(i + slideindex, 2, case)
        ws.cell(i + slideindex, 3, origfile)
        ws.cell(i + slideindex, 4, anonid)
        ws.cell(i + slideindex, 5, block)
        ws.cell(i + slideindex, 6, stain)
        if slideindex > 0:
            for c in range(7, ws.max_column):
                ws.cell(i + slideindex, c, ws.cell(i, c).value)
        mark_done(ws, i + slideindex, case, origfile)

def get_str(worksheet, row, column):
    return str(worksheet.cell(row, column).value or '').strip()

def get_caseid(case):
    if case.endswith('.zip'):
        return case[:-4]
    return case

def parse_anonid(i, anonid):
    try:
        digits = len(re.search(r'\d+$', anonid)[0])
    except:
        err(i, "AnonID {!r} must end with a zero padded number, such as '001'.",
            anonid)
    prefix = anonid[:-digits]
    return prefix, digits

def validate_anonymization_data(worksheet):
    anonids = dict() # caseid -> anonid
    caseids = dict() # anonid -> caseid
    caseid_rows = dict() # caseid -> rownumber
    barcode_rows = dict() # barcode -> rownumber
    done = set()
    prefix = None
    digits = 0
    anonid_number = 0
    for i in range(13, worksheet.max_row + 1):
        status = get_str(worksheet, i, 1)
        caseid = get_caseid(os.path.basename(get_str(worksheet, i, 2)))
        origfile = get_str(worksheet, i, 3)
        anonid = get_str(worksheet, i, 4)
        block = get_str(worksheet, i, 5)
        stain = get_str(worksheet, i, 6)
        if status:
            if status.lower() == 'ignore':
                continue
            if status.lower() != "done":
                err(i, "Unknown Status {!r}.", status)
            if not (anonid and block and stain):
                err(i, "AnonID/Block/Stain missing!")
            barcode = get_barcode(anonid, block, stain)
            if barcode in barcode_rows:
                err(i, "Barcode {!r} same as on Row {} but bacodes must be unique.",
                    barcode, barcode_rows[barcode])
            barcode_rows[barcode] = i
        if not anonid:
            err(i, "No AnonID given.")
        if prefix is None:
            prefix, digits = parse_anonid(i, anonid)
        anonid_number = validate_anonid_number(i, anonid, prefix, digits, anonid_number)
        if caseid:
            validate_id_mapping(i, caseid, anonid, caseids, anonids)
            if caseid in caseid_rows and not (status.lower() == 'done' and caseid in done):
                err(i, "Cases must be unique before anonymizing, but ID {!r} occurs multiple times; here and also on row {}.",
                    caseid, caseid_rows[caseid])
            if status.lower() == 'done':
                done.add(caseid)
            elif origfile or block or stain:
                err(i, "Garbage in OrigFile/Block/Stain columns; OrigFile, Block and Stain must be given by subfolders to Case.")
            caseid_rows[caseid] = i
        elif status.lower() != "done":
            err(i, "No Case specified.")
    return set(barcode_rows.keys())

def anonymize(wb, basedir, tmpdir, anondir, excelfile):
    ws = wb.active
    check_worksheet(ws)
    barcodes = validate_anonymization_data(ws)

    i = 13
    while i <= ws.max_row:
        # Status	Case	OrigFile	AnonID	Block	Stain
        status = get_str(ws, i, 1)
        case = os.path.basename(get_str(ws, i, 2))
        caseid = get_caseid(case)
        is_zip = case.endswith('.zip')
        origfile = get_str(ws, i, 3)
        anonid = get_str(ws, i, 4)
        block = get_str(ws, i, 5)
        stain = get_str(ws, i, 6)

        if status:
            if status.lower() == "done":
                mark_done(ws, i, case, origfile)
                wb.save(excelfile)
            i += 1
            continue

        casefile = os.path.join(basedir, case)
        if not os.path.exists(casefile):
            print("Terminating at Row {}: Case {!r} does not exist in work directory {}.".format(
                i, case, basedir))
            return barcodes

        casedir = casefile
        if is_zip:
            casedir = os.path.join(tmpdir, caseid)
            print("Decompressing {}...".format(case))
            zipfile.ZipFile(casefile).extractall(tmpdir)

        slides = get_slides(i, casedir)
        if not slides:
            err(i, "No slides present for Case {!r}.", case)
        for block, stain, origfile in slides:
            barcode = get_barcode(anonid, block, stain)
            anonymize_slide(i, origfile, anondir, barcode)
            barcodes.add(barcode)

        update_spreadsheet(i, ws, case, origfile, anonid, slides)
        wb.save(excelfile)
        i += len(slides)

        if is_zip:
            shutil.rmtree(casedir)

    return barcodes

def get_garbage(anondir, barcodes):
    garbage = []
    for filename in files(anondir):
        barcode = filename.rsplit('_', 1)[0]
        if barcode not in barcodes:
            garbage.append(filename)
    return garbage

def main(argv):
    options = get_options(argv)
    wb = openpyxl.load_workbook(options.excelfile)
    base, ext = os.path.splitext(options.excelfile.name)
    anonexcelfile = base + '_anon' + ext
    basedir = os.path.dirname(options.excelfile.name)
    if not options.tmpdir:
        options.tmpdir = os.path.join(basedir, base + '_tmp')
    if not os.path.isdir(options.tmpdir):
        os.makedirs(options.tmpdir, 0o755)
    if not options.anondir:
        options.anondir = os.path.join(basedir, base + '_anon')
    if not os.path.isdir(options.anondir):
        os.makedirs(options.anondir, 0o755)
    try:
        barcodes = anonymize(wb, basedir, options.tmpdir, options.anondir, options.excelfile.name)
    except ParseError as err:
        print("Error on row {}: {}".format(*err.args), file=sys.stderr)
        if options.suppress_exitcode:
            return None
        return 1
    garbage = get_garbage(options.anondir, barcodes)
    if garbage:
        print("\nWarning: Possible garbage files found in {!r} folder. You may want to delete these before proceding:".format(options.anondir), file=sys.stderr)
        print("\n".join(repr(s) for s in garbage), file=sys.stderr)
    print("\nDone! Anonymized images in {!r} folder. {!r} has been updated.".format(
            os.path.basename(options.anondir), os.path.basename(options.excelfile.name)))
    print("\nYour data is now Pseudonymous.\n")
    print("To make your data Anonymous: Delete all keys associating AnonIDs to "
          "Cases, including the Case and OrigFile file cells in {0!r}. "
          "Obviously, none of the study parameters in {0!r} may contain "
          "identifiers either.".format(os.path.basename(options.excelfile.name)))

if __name__ == "__main__":
    exitcode = main(sys.argv)
    if exitcode is not None:
        sys.exit(exitcode)
